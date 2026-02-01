"""
엑셀 저장 모듈
수집한 데이터를 엑셀 파일로 저장합니다.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# 엑셀 저장 시 사용할 컬럼 순서 (왼쪽부터): index, 숙소명, 위치, 가격, 가격상세, 방구성, 상세링크, 숙소ID, 할인정보
EXCEL_COLUMNS_ORDER = [
    'index', 'title', 'location', 'price', 'price_detail',
    'room_composition', 'url', 'room_id', 'discount'
]

# 컬럼명 매핑 (영문 → 한글)
COLUMN_MAPPING = {
    'index': 'Index',
    'title': '위치',
    'price': '가격',
    'location': '숙소명',
    'room_composition': '방구성',
    'price_detail': '가격상세',
    'discount': '할인정보',
    'url': '상세링크',
    'room_id': '숙소ID',
}


def generate_filename(prefix: str = "33m2_숙소목록") -> str:
    """
    타임스탬프가 포함된 파일명 생성

    Args:
        prefix: 파일명 접두사

    Returns:
        생성된 파일명 (확장자 포함)
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"


def save_to_excel(data: list, filepath: str) -> bool:
    """
    크롤링한 데이터를 엑셀 파일로 저장

    Args:
        data: 크롤링한 데이터 리스트
        filepath: 저장할 파일 경로

    Returns:
        bool: 저장 성공 여부
    """
    if not data:
        logger.warning("저장할 데이터가 없습니다.")
        return False

    try:
        # DataFrame 생성
        df = pd.DataFrame(data)

        # 지정한 순서의 컬럼만 유지 (없는 컬럼은 빈 값으로 추가)
        for col in EXCEL_COLUMNS_ORDER:
            if col not in df.columns:
                df[col] = ''
        df = df[EXCEL_COLUMNS_ORDER]

        # 컬럼명을 한글로 변환
        df = df.rename(columns=COLUMN_MAPPING)

        # 리스트/딕셔너리 타입을 문자열로 변환
        for col in df.columns:
            df[col] = df[col].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
            df[col] = df[col].apply(lambda x: str(x) if isinstance(x, dict) else x)

        # 엑셀 파일로 저장
        df.to_excel(filepath, index=False, engine='openpyxl')

        # 서식 적용
        _apply_formatting(filepath)

        logger.info(f"엑셀 파일 저장 완료: {filepath}")
        return True

    except Exception as e:
        logger.error(f"엑셀 저장 실패: {str(e)}")
        return False


def _apply_formatting(filepath: str) -> None:
    """
    엑셀 파일에 서식 적용

    Args:
        filepath: 엑셀 파일 경로
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        # 헤더 스타일
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 셀 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 행 서식 적용
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 행 서식 적용
        data_alignment = Alignment(vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

        # 자동 열 너비 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    # 한글은 2배 너비로 계산
                    cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass

            # 최대/최소 너비 제한
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 첫 번째 행 고정
        ws.freeze_panes = 'A2'

        wb.save(filepath)

    except Exception as e:
        logger.warning(f"서식 적용 중 오류: {str(e)}")


def get_default_save_path() -> str:
    """
    기본 저장 경로 반환 (실행파일/스크립트와 동일한 위치)
    """
    from .config import DATA_DIR
    return str(DATA_DIR / generate_filename())


def export_to_csv(data: list, filepath: str) -> bool:
    """
    데이터를 CSV 파일로 저장 (백업용)

    Args:
        data: 크롤링한 데이터 리스트
        filepath: 저장할 파일 경로

    Returns:
        bool: 저장 성공 여부
    """
    if not data:
        logger.warning("저장할 데이터가 없습니다.")
        return False

    try:
        df = pd.DataFrame(data)
        for col in EXCEL_COLUMNS_ORDER:
            if col not in df.columns:
                df[col] = ''
        df = df[EXCEL_COLUMNS_ORDER]
        df = df.rename(columns=COLUMN_MAPPING)

        # 리스트/딕셔너리를 문자열로 변환
        for col in df.columns:
            df[col] = df[col].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
            df[col] = df[col].apply(lambda x: str(x) if isinstance(x, dict) else x)

        csv_path = filepath.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')

        logger.info(f"CSV 파일 저장 완료: {csv_path}")
        return True

    except Exception as e:
        logger.error(f"CSV 저장 실패: {str(e)}")
        return False
